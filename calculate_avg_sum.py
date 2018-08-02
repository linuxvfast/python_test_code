import openpyxl

def process_worksheet(sheet):
    sum_column = sheet.max_column + 1
    avg_column = sheet.max_column + 2

    for row in sheet.iter_rows(min_row=2,min_col=3):  #iter_rows按行获取每一行的数据
        scores = [cell.value for cell in row]
        sum_score = sum(scores)
        avg_score = sum_score /len(scores)
        print(row[0])
        print(row[0].row)  #获取单元格所在的行
        sheet.cell(row=row[0].row,column=sum_column).value = sum_score
        sheet.cell(row=row[0].row, column=avg_column).value = avg_score

    #设置标题
    sheet.cell(row=1, column=sum_column).value = 'sum'
    sheet.cell(row=1, column=avg_column).value = 'avg'




def main():
    workboot = openpyxl.load_workbook('example.xlsx')
    sheet = workboot.get_sheet_by_name('student')
    process_worksheet(sheet)
    workboot.save('example_copy.xlsx')


if __name__ == '__main__':
    main()
