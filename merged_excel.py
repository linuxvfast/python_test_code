import os
import glob
import openpyxl

#合并excel表格
def merge_xlsx_file(xlsx_files):
    #修改第一个活动表的标题
    workboot = openpyxl.load_workbook(xlsx_files[0])
    worksheet = workboot.active
    worksheet.title = "merged result"
    #在表中追加其它表中的数据
    for filename in xlsx_files[1:]:
        boot = openpyxl.load_workbook(filename)
        sheet = boot.active
        for row in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            worksheet.append(values)

    return workboot


#查询所有的excel表
def get_all_xlsx_files(path):
    xlsx_files = glob.glob(os.path.join(path,'*.xlsx'))
    sorted(xlsx_files,key=str.lower)
    return xlsx_files

def main():
    os.chdir('excel')
    path = os.getcwd()
    print(path)
    xlsx_files = get_all_xlsx_files(path)
    workboot = merge_xlsx_file(xlsx_files)
    workboot.save('merged_form.xlsx')



if __name__ == '__main__':
    main()
